"""DORMANT — interim workbook renderer, superseded by the /state-of-play page.

The E01 platform module (lib/recruiting-ops/modules/exec-state-of-play.ts) is
now the canonical facts producer; its data pull sidecar
(recruiting-ops-exec-sheet-data.ts) is retired. This renderer is retained only
as a candidate for the future C4 sheets-export lane; it is not invoked by any
runner and its health/idle logic does NOT match the platform's governed
definitions. Do not hand-run it for stakeholders.

Executive Open-Requisitions workbook (v2 rebuild) from exec-sheet-data.json.

Design contract (per rebuild brief v2):
  - ALL open reqs org-wide. Every headline number is a formula over the same
    named Table the reader sees below it, so cards reconcile by construction.
  - No KPI-card strip (reads as generic): a proportional health-composition bar
    + an understated inline stat line instead.
  - Live: named Excel Tables (tblRoles, tblFunnel, tblOffers, tblData) with
    in-table structured-reference formulas for every derived value (idle, health,
    offers·12wk, status). Health COLOR comes only from conditional-formatting
    rules keyed on the health value. Req numeric, dates real.
  - Full Greenhouse stage names verbatim everywhere; funnel headers rotated 90°.
    The full stage funnel is its own 1-page-wide tab — 21 readable columns can't
    honestly fit one landscape page without breaking the 9pt floor.
  - Candidate references are hyperlinked names (=HYPERLINK to the GH profile);
    zero raw IDs on exec sheets.
  - Each exec tab: landscape, fit 1 page wide, repeating header, print area.

Usage: python3 scripts/build-exec-workbook.py <exec-sheet-data.json> <out.xlsx>
"""
import json
import sys
from datetime import date, datetime

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.properties import PageSetupProperties

INK, MUTED, FAINT, HAIR, SLATE = "1F2933", "7B8794", "B8C1CC", "E4E7EB", "24344D"
CARD = "F5F7FA"
GREEN, AMBER, RED = "2E7D46", "B7791F", "C0392B"
FONT = "Aptos"  # modern default; falls back cleanly
THIN = Side(style="thin", color=HAIR)


def F(size=11, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def health_of(r):
    active, seats, idle = r["active"], r["seats"], r["idleDays"]
    if active == 0 and seats > 0:
        return "Red"
    if idle is None:
        return "Green"
    if idle > 14:
        return "Red"
    if idle > 7 or active < 3:
        return "Amber"
    return "Green"


HRANK = {"Red": 0, "Amber": 1, "Green": 2}


def sorted_roles(roles):
    # departments ordered by attention (reds, then ambers), then within dept red→amber→green→idle desc
    dept_attn = {}
    for r in roles:
        d = r["department"]
        red = 1 if r["_health"] == "Red" else 0
        amb = 1 if r["_health"] == "Amber" else 0
        cur = dept_attn.get(d, [0, 0])
        dept_attn[d] = [cur[0] + red, cur[1] + amb]
    dept_order = {d: i for i, (d, _) in enumerate(
        sorted(dept_attn.items(), key=lambda kv: (-kv[1][0], -kv[1][1], kv[0]))
    )}
    return sorted(roles, key=lambda r: (dept_order[r["department"]], HRANK[r["_health"]], -(r["idleDays"] or 0)))


def style_title(ws, first_col, span, title, subtitle):
    ws.merge_cells(start_row=2, start_column=first_col, end_row=2, end_column=first_col + span - 1)
    t = ws.cell(2, first_col, title); t.font = F(18, bold=True); t.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 26
    ws.merge_cells(start_row=3, start_column=first_col, end_row=3, end_column=first_col + span - 1)
    s = ws.cell(3, first_col, subtitle); s.font = F(11, color=MUTED)


def header_row(ws, row, start_col, cols, rotate_from=None):
    for i, (label, width, align) in enumerate(cols):
        c = start_col + i
        ws.column_dimensions[get_column_letter(c)].width = width
        cell = ws.cell(row, c, label)
        cell.fill = PatternFill("solid", fgColor=SLATE)
        rot = 90 if (rotate_from is not None and i >= rotate_from) else 0
        cell.font = F(10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal=("center" if align == "c" else "left"),
                                   vertical=("bottom" if rot else "center"), wrap_text=True, text_rotation=rot)
    ws.row_dimensions[row].height = 96 if rotate_from is not None else 24


def add_table(ws, name, top, left, nrows, ncols):
    ref = f"{get_column_letter(left)}{top}:{get_column_letter(left + ncols - 1)}{top + nrows}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False, showColumnStripes=False,
                                      showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)


def landscape(ws, print_ref, title_rows):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = print_ref
    ws.print_title_rows = title_rows
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4


def hairline_rows(ws, first, last, left, right):
    for r in range(first, last + 1):
        for c in range(left, right + 1):
            ws.cell(r, c).border = Border(bottom=THIN)


def finalist_formula(finalists):
    if not finalists:
        return None
    top = finalists[0]
    label = f"{top['name']} ({top['stage']})"
    extra = len(finalists) - 1
    if extra:
        label += f"  +{extra} more"
    url = top["url"] or ""
    return f'=HYPERLINK("{url}","{label.replace(chr(34), chr(39))}")'


# ============================================================ Open Roles
def build_open_roles(wb, roles, funnel_stages, n_off):
    ws = wb.active
    ws.title = "Open Roles"
    ws.sheet_properties.tabColor = SLATE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ncols_for_span = 12
    style_title(ws, 2, ncols_for_span, "Recruiting — Open Requisitions",
                "All open reqs, org-wide · grouped by department, red → amber → green within · trailing-12-week offers")

    reds = sum(1 for r in roles if r["_health"] == "Red")
    ambers = sum(1 for r in roles if r["_health"] == "Amber")
    greens = sum(1 for r in roles if r["_health"] == "Green")
    total = len(roles)

    # Health composition bar (rich text, proportional block glyphs) — the hero,
    # not a card strip. ~48 blocks total, split by health.
    def blocks(n):
        return "▉" * max(1, round(48 * n / total)) if n else ""
    bar = CellRichText()
    if reds:
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=13, color=RED), blocks(reds)))
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=11, b=True, color=RED), f" {reds} red   "))
    if ambers:
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=13, color=AMBER), blocks(ambers)))
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=11, b=True, color=AMBER), f" {ambers} amber   "))
    if greens:
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=13, color=GREEN), blocks(greens)))
        bar.append(TextBlock(InlineFont(rFont=FONT, sz=11, b=True, color=GREEN), f" {greens} on track"))
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=13)
    bc = ws.cell(5, 2); bc.value = bar; bc.alignment = Alignment(vertical="center")
    ws.row_dimensions[5].height = 24

    # One live summary line (single merged cell, no card strip) — reconciles by
    # reading the same ranges the table below shows. Row math known ahead: header
    # at row 11, data rows 12..(11+n).
    n = len(roles)
    ro_first, ro_last = 12, 11 + n
    off_last = 4 + n_off
    summary = (f'=COUNTA(C{ro_first}:C{ro_last})&" open reqs org-wide      "'
               f'&(COUNTIF(N{ro_first}:N{ro_last},"Red")+COUNTIF(N{ro_first}:N{ro_last},"Amber"))&" need attention (red + amber)      "'
               f'&SUM(J{ro_first}:J{ro_last})&" seats to fill      "'
               f'&COUNTA(Offers!B5:B{off_last})&" offers accepted, trailing 12 weeks"')
    sc = ws.cell(7, 2, summary); sc.font = F(15, color=INK); sc.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=13)
    ws.row_dimensions[7].height = 24
    ws.cell(8, 2, "Seats = positions to fill; one req can carry several.  ·  Idle = days since the most recent candidate activity on the req.  ·  Health from idle-day thresholds.").font = F(9, color=MUTED)
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=13)

    hrow = 11
    cols = [
        ("●", 3, "c"), ("Role", 34, "l"), ("Req", 7, "c"), ("Department", 22, "l"), ("Owner", 18, "l"),
        ("Stage", 22, "l"), ("Active", 8, "c"), ("Idle days", 9, "c"), ("Seats", 7, "c"),
        ("Offers · 12 wk", 9, "c"), ("Finalists (Assessment and later)", 40, "l"), ("Status note", 40, "l"),
        ("Health", 9, "l"), ("Last activity", 12, "c"),
    ]
    header_row(ws, hrow, 2, cols)
    ordered = sorted_roles(roles)
    r0 = hrow + 1
    for i, role in enumerate(ordered):
        r = r0 + i
        ws.cell(r, 2, "●").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 3, role["role"]).font = F(11, bold=True)
        rq = ws.cell(r, 4, role["reqId"]); rq.alignment = Alignment(horizontal="center", vertical="center"); rq.font = F(11, color=MUTED)
        ws.cell(r, 5, role["department"]).font = F(10)
        ws.cell(r, 6, role["owner"] or "—").font = F(10, color=(INK if role["owner"] else RED))
        ws.cell(r, 7, role["furthestStage"]).font = F(11)
        av = ws.cell(r, 8, role["active"]); av.alignment = Alignment(horizontal="center", vertical="center"); av.font = F(11)
        idle = ws.cell(r, 9, f'=IF(H{r}=0,"",TODAY()-O{r})')
        idle.alignment = Alignment(horizontal="center", vertical="center"); idle.number_format = '0;;'
        sv = ws.cell(r, 10, role["seats"]); sv.alignment = Alignment(horizontal="center", vertical="center"); sv.font = F(11)
        ov = ws.cell(r, 11, f'=COUNTIFS(Offers!$D$5:$D${off_last},D{r},Offers!$F$5:$F${off_last},">="&(TODAY()-84))')
        ov.alignment = Alignment(horizontal="center", vertical="center"); ov.number_format = '0;;'
        fin = finalist_formula(role["finalists"])
        fc = ws.cell(r, 12, fin if fin else "—")
        fc.font = F(10, color=("0563C1" if fin else MUTED)); fc.alignment = Alignment(vertical="center")
        st = ws.cell(r, 13, f'=IF(N{r}="Green","",IF(H{r}=0,"No active candidates against "&J{r}&" open seat(s)",'
                            f'IF(I{r}>14,"Stalled — no activity in "&I{r}&" days",'
                            f'IF(I{r}>7,"Slowing — "&I{r}&" days idle","Thin pipeline — "&H{r}&" active"))))')
        st.font = F(10, color=MUTED); st.alignment = Alignment(vertical="center", wrap_text=True)
        hc = ws.cell(r, 14, f'=IF(AND(H{r}=0,J{r}>0),"Red",IF(O{r}="","Green",'
                            f'IF(I{r}>14,"Red",IF(OR(I{r}>7,H{r}<3),"Amber","Green"))))')
        hc.font = F(10, color=MUTED)
        la = ws.cell(r, 15, role["lastActivity"] and parse_date(role["lastActivity"]))
        la.number_format = "d-mmm"; la.alignment = Alignment(horizontal="center", vertical="center"); la.font = F(10, color=MUTED)
        ws.row_dimensions[r].height = 26
    last = r0 + len(ordered) - 1

    add_table(ws, "tblRoles", hrow, 2, len(ordered), len(cols))
    # dot color via CF keyed on Health (col N = index 14)
    dot_range = f"B{r0}:B{last}"
    hcol = get_column_letter(14)
    for word, color in (("Red", RED), ("Amber", AMBER), ("Green", GREEN)):
        ws.conditional_formatting.add(dot_range, FormulaRule(formula=[f'${hcol}{r0}="{word}"'],
                                                             font=Font(name=FONT, size=13, color=color)))
    hairline_rows(ws, r0, last, 2, 13)
    # hide helper columns
    ws.column_dimensions["N"].hidden = True
    ws.column_dimensions["O"].hidden = True
    ws.freeze_panes = ws.cell(r0, 4)
    landscape(ws, f"B2:M{last}", f"${hrow}:${hrow}")


# ============================================================ Pipeline by Stage
def build_funnel(wb, roles, funnel_stages):
    ws = wb.create_sheet("Pipeline by Stage")
    ws.sheet_properties.tabColor = SLATE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    style_title(ws, 2, 6, "Pipeline by Stage",
                "Active candidates by interview stage · full Greenhouse stage names · one row per open req")
    hrow = 4
    cols = [("●", 3, "c"), ("Role", 34, "l"), ("Req", 7, "c"), ("Department", 20, "l")]
    cols += [(s, 4.4, "c") for s in funnel_stages]
    cols += [("Active", 8, "c")]
    header_row(ws, hrow, 2, cols, rotate_from=4)
    # health helper header (last table column, hidden) — drives the dot CF
    hh = ws.cell(hrow, 6 + len(funnel_stages) + 1, "Health")
    hh.fill = PatternFill("solid", fgColor=SLATE); hh.font = F(10, bold=True, color="FFFFFF")
    ordered = sorted_roles(roles)
    r0 = hrow + 1
    nfun = len(funnel_stages)
    for i, role in enumerate(ordered):
        r = r0 + i
        ws.cell(r, 2, "●").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 3, role["role"]).font = F(11, bold=True)
        rq = ws.cell(r, 4, role["reqId"]); rq.alignment = Alignment(horizontal="center"); rq.font = F(11, color=MUTED)
        ws.cell(r, 5, role["department"]).font = F(10)
        for j, s in enumerate(funnel_stages):
            v = role["funnel"].get(s, 0)
            c = ws.cell(r, 6 + j, v if v else None)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = F(10, color=(INK if v else FAINT))
            c.number_format = '0;;'
        actc = 6 + nfun
        a = ws.cell(r, actc, f"=SUM({get_column_letter(6)}{r}:{get_column_letter(6+nfun-1)}{r})")
        a.alignment = Alignment(horizontal="center", vertical="center"); a.font = F(11, bold=True)
        # health helper (hidden col after Active) for dot CF
        hcol_idx = actc + 1
        ws.cell(r, hcol_idx, role["_health"]).font = F(9, color=MUTED)
        ws.row_dimensions[r].height = 22
    last = r0 + len(ordered) - 1
    total_cols = 4 + nfun + 1
    add_table(ws, "tblFunnel", hrow, 2, len(ordered), total_cols + 1)  # +1 health helper
    hcol = get_column_letter(6 + nfun + 1)
    for word, color in (("Red", RED), ("Amber", AMBER), ("Green", GREEN)):
        ws.conditional_formatting.add(f"B{r0}:B{last}", FormulaRule(formula=[f'${hcol}{r0}="{word}"'],
                                                                    font=Font(name=FONT, size=13, color=color)))
    ws.column_dimensions[get_column_letter(6 + nfun + 1)].hidden = True
    hairline_rows(ws, r0, last, 2, 5 + nfun)
    ws.freeze_panes = ws.cell(r0, 5)
    landscape(ws, f"B2:{get_column_letter(6+nfun)}{last}", f"${hrow}:${hrow}")


# ============================================================ ELT Briefing
def build_elt(wb, roles, offers):
    ws = wb.create_sheet("ELT Briefing")
    ws.sheet_properties.tabColor = SLATE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    style_title(ws, 2, 6, "ELT Briefing — Recruiting",
                "Per-department rollup over the open-req portfolio · trailing-12-week offers")
    depts = sorted({r["department"] for r in roles})
    nr = len(roles)
    ro1, ro2 = 12, 11 + nr          # Open Roles data rows
    off2 = 4 + len(offers)          # Offers data last row
    RO = "'Open Roles'"
    hrow = 4
    cols = [("Department", 44, "l"), ("Open reqs", 10, "c"), ("Needs attention (red + amber)", 16, "c"),
            ("Offers · 12 wk", 12, "c"), ("Seats", 8, "c")]
    header_row(ws, hrow, 2, cols)
    ws.row_dimensions[hrow].height = 40
    r = hrow + 1
    for d in depts:
        dd = d.replace('"', "'")
        formulas = {
            2: d,
            3: f'=COUNTIF({RO}!$E${ro1}:$E${ro2},"{dd}")',
            4: f'=COUNTIFS({RO}!$E${ro1}:$E${ro2},"{dd}",{RO}!$N${ro1}:$N${ro2},"Red")+COUNTIFS({RO}!$E${ro1}:$E${ro2},"{dd}",{RO}!$N${ro1}:$N${ro2},"Amber")',
            5: f'=COUNTIF(Offers!$E$5:$E${off2},"{dd}")',
            6: f'=SUMIF({RO}!$E${ro1}:$E${ro2},"{dd}",{RO}!$J${ro1}:$J${ro2})',
        }
        for c, val in formulas.items():
            cell = ws.cell(r, c, val)
            cell.font = F(11, bold=(c == 2))
            if c > 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1
    add_table(ws, "tblDept", hrow, 2, len(depts), len(cols))
    hairline_rows(ws, hrow + 1, r - 1, 2, 6)

    # Narrative built by formula — labeled lines, not a blob.
    nr = r + 2
    ws.cell(nr, 2, "NARRATIVE (DRAFT)").font = F(9, bold=True, color=MUTED); nr += 1
    E, N, F_owner = f"{RO}!$E${ro1}:$E${ro2}", f"{RO}!$N${ro1}:$N${ro2}", f"{RO}!$F${ro1}:$F${ro2}"
    C_role = f"{RO}!$C${ro1}:$C${ro2}"
    offC = f"Offers!$B$5:$B${off2}"
    lines = [
        ('Headline', f'=COUNTA({C_role})&" open reqs across "&{len(depts)}&" departments; "&(COUNTIF({N},"Red")+COUNTIF({N},"Amber"))&" need attention; "&COUNTA({offC})&" offers accepted in the trailing 12 weeks."'),
        ('Risks', f'=IF(COUNTIF({N},"Red")=0,"No red reqs.",COUNTIF({N},"Red")&" red reqs — open seats with stalled or empty pipelines. See the Open Roles tab, sorted red-first.")'),
        ('Asks', f'=IF(COUNTBLANK({F_owner})=0,"All open reqs have an owner.",COUNTIF({F_owner},"—")&" open reqs have no recruiter/sourcer owner — assign before they stall.")'),
    ]
    for label, formula in lines:
        ws.cell(nr, 2, label).font = F(11, bold=True)
        cell = ws.cell(nr, 3, formula); cell.font = F(11); cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=nr, start_column=3, end_row=nr, end_column=6)
        ws.row_dimensions[nr].height = 52
        nr += 1
    landscape(ws, f"B2:F{nr-1}", f"${hrow}:${hrow}")


# ============================================================ Offers
def build_offers(wb, offers):
    ws = wb.create_sheet("Offers")
    ws.sheet_properties.tabColor = SLATE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    style_title(ws, 2, 5, "Offers Accepted — Trailing 12 Weeks",
                "One row per accepted offer · candidate links to Greenhouse")
    hrow = 4
    cols = [("Candidate", 30, "l"), ("Role", 38, "l"), ("Req", 7, "c"),
            ("Department", 24, "l"), ("Accepted", 12, "c"), ("Starts", 12, "c")]
    header_row(ws, hrow, 2, cols)
    r0 = hrow + 1
    for i, o in enumerate(offers):
        r = r0 + i
        name = o["candidate"].replace('"', "'")
        link = f'=HYPERLINK("{o["url"]}","{name}")' if o["url"] else name
        cc = ws.cell(r, 2, link); cc.font = F(11, color=("0563C1" if o["url"] else INK))
        ws.cell(r, 3, o["role"]).font = F(11)
        rq = ws.cell(r, 4, o["reqId"]); rq.alignment = Alignment(horizontal="center"); rq.font = F(11, color=MUTED)
        ws.cell(r, 5, o["department"]).font = F(10)
        ac = ws.cell(r, 6, parse_date(o["acceptedAt"])); ac.number_format = "d-mmm"; ac.alignment = Alignment(horizontal="center")
        sc = ws.cell(r, 7, parse_date(o["startsOn"])); sc.number_format = "d-mmm"; sc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22
    last = r0 + len(offers) - 1
    add_table(ws, "tblOffers", hrow, 2, len(offers), len(cols))
    hairline_rows(ws, r0, last, 2, 7)
    ws.freeze_panes = ws.cell(r0, 2)
    landscape(ws, f"B2:G{last}", f"${hrow}:${hrow}")


# ============================================================ Data (source)
def build_data(wb, roles, funnel_stages):
    ws = wb.create_sheet("Data")
    ws.sheet_properties.tabColor = MUTED
    ws.sheet_view.showGridLines = False
    headers = ["Req", "Role", "Department", "Owner", "Opened", "Seats", "Active",
               "Last activity", "Idle days", "Furthest stage", "Health", "Finalist count"] + list(funnel_stages)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.font = F(10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    ordered = sorted_roles(roles)
    for i, role in enumerate(ordered):
        r = 2 + i
        vals = [role["reqId"], role["role"], role["department"], role["owner"],
                parse_date(role["openedAt"]), role["seats"], role["active"],
                parse_date(role["lastActivity"]), role["idleDays"], role["furthestStage"],
                role["_health"], len(role["finalists"])]
        vals += [role["funnel"].get(s, 0) for s in funnel_stages]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v); cell.font = F(10)
            if headers[j-1] in ("Opened", "Last activity"):
                cell.number_format = "d-mmm-yy"
    add_table(ws, "tblData", 1, 1, len(ordered), len(headers))
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.freeze_panes = "A2"
    # Source tab — compress its print footprint so it doesn't flood the PDF.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = f"A1:{get_column_letter(len(headers))}{1+len(ordered)}"
    ws.print_title_rows = "$1:$1"


def main():
    src, out = sys.argv[1], sys.argv[2]
    data = json.load(open(src))
    roles = data["roles"]
    funnel_stages = data["funnelStages"]
    offers = data["offers"]
    for r in roles:
        r["_health"] = health_of(r)
    wb = openpyxl.Workbook()
    build_open_roles(wb, roles, funnel_stages, len(offers))
    build_funnel(wb, roles, funnel_stages)
    build_elt(wb, roles, offers)
    build_offers(wb, offers)
    build_data(wb, roles, funnel_stages)
    wb.save(out)
    reds = sum(1 for r in roles if health_of(r) == "Red")
    ambers = sum(1 for r in roles if health_of(r) == "Amber")
    print(f"wrote {out}: {len(roles)} reqs ({reds} red, {ambers} amber, {len(roles)-reds-ambers} green), "
          f"{len(offers)} offers, tabs={wb.sheetnames}")


if __name__ == "__main__":
    main()
